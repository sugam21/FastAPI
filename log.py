import datetime
from pathlib import Path

import openpyxl
import pandas as pd
from loguru import logger


class Log:
    def __init__(self):
        """Check if file exists. IF not create one if yes do nothing"""
        self.log_save_file: Path = Path(".").resolve() / "data" / "logs" / "logs.xlsx"
        self._check_folder()
        self._check_file()

    def _check_folder(self):
        if self.log_save_file.parent.exists():
            logger.info("Log directory already exists.")
        else:
            logger.warning(
                f"Log directory is missing. Creating a new one in {self.log_save_file.parent}."
            )
            self.log_save_file.parent.mkdir(parents=True, exist_ok=True)

    def _check_file(self):
        if (self.log_save_file).is_file():
            logger.info("Log File already exists.")
        else:
            logger.warning(
                f"Log File does not exists. Creating a new one in {self.log_save_file}"
            )

            wb = openpyxl.Workbook()
            sheets: list[str] = ["Accounts", "Policies", "Claims"]
            for idx, sheet in enumerate(sheets):
                wb.create_sheet(sheet, index=idx)
                ws = wb[sheet]
                match sheet:
                    case "Accounts":
                        ws.append(
                            ["Timestamp", "AccountId", "Column", "OldValue", "NewValue"]
                        )
                    case "Policies":
                        ws.append(
                            ["Timestamp", "HAN", "Column", "OldValue", "NewValue"]
                        )
                    case "Claims":
                        ws.append(["Timestamp", "Id", "Column", "OldValue", "NewValue"])

            wb.save(self.log_save_file)

    def _write(self, sheet_name, new_row):
        try:
            wb = openpyxl.load_workbook(self.log_save_file)
        except FileNotFoundError as e:
            logger.error(f"The Log file not found: {e}")
        else:
            ws = wb[sheet_name]
            ws.append(new_row)
            try:
                wb.save(self.log_save_file)
            except PermissionError:
                logger.error(
                    "The file is open in another program. Close the file and try again."
                )
            else:
                logger.info(f"Log file for sheet {sheet_name} written successfully")

    def write_log(
        self, sheet_name: str, id: str, column_name: str, old_value: any, new_value: any
    ) -> dict:
        """Check if the sheet exists or not. If not then creates it. If it does then appends it."""

        valid_sheets: list[str] = ["Accounts", "Claims", "Policies"]
        if sheet_name not in valid_sheets:
            logger.error(f"Sheet name {sheet_name} not valid.")

        new_row: tuple = (
            datetime.datetime.now(),
            id,
            column_name,
            old_value,
            new_value,
        )
        self._write(sheet_name=sheet_name, new_row=new_row)


if __name__ == "__main__":
    obj = Log()
    obj.write_log(
        sheet_name="Accounts",
        id="12345",
        column_name="acc1",
        old_value=123,
        new_value=12345,
    )
    obj.write_log(
        sheet_name="Claims",
        id="12345",
        column_name="acc1",
        old_value=123,
        new_value=12345,
    )
    obj.write_log(
        sheet_name="Policies",
        id="12345",
        column_name="acc1",
        old_value=123,
        new_value=12345,
    )
